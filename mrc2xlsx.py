# import openpyxl as opx
from openpyxl import workbook
from openpyxl.styles import Font, colors, Alignment
import sys

SUBFIELD_PREFIX = "$$"
HEADER = ["Kat.", "Ind.", "SF", "Feldinhalt"]


# Namen der Input- und Output-Dateien
#if len(argv) < 2:
#    in_file_name = input("Bitte geben Sie den Dateinamen der Quelldatei ein: \n")
#else:
#    in_file_name = argv[1]

in_file_name = "TEST.MRC"

out_file_name = in_file_name[:-4] + ".xlsx"

# xlsx-Workbook
wb = opx.Workbook()
ws = wb.active
ws.title = in_file_name[:-4]
ws.append(HEADER)

kat_nr = ws.column_dimensions['A']
ind = ws.column_dimensions['B']
subfield = ws.column_dimensions['C']
sf_data = ws.column_dimensions['D']
head_row = ws.row_dimensions[1]

# styles für xlsx-output h
head_row.font = Font(name='Calibri',
                    size=11,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False)

kat_nr.font = Font(name='Courier New',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='single',
                    strike=False,
                    color=colors.BLUE)
kat_nr.alignment = Alignment(vertical='top')

ind.font = Font(name='Courier New',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='single',
                    strike=False,
                    color=colors.BLUE)
ind.alignment = Alignment(vertical='top')

subfield.font = Font(name='Courier New',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='single',
                    strike=False,
                    color=colors.RED)
subfield.alignment = Alignment(shrink_to_fit=True)

sf_data.font = Font(name='Courier New',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False)

 
# open input file line by line into a list
mrc_in = open(in_file_name).read().splitlines()

def process_line(line):
    """
    Arbeitet Zeile für Zeile ab
    """

    out = []
    
    # Liste der Werte in Subfeldern erstellen
    sf_vals = []
    if "$" in line:
        sf_vals = line[10:].split("$$")
    else:
        # für Felder ohne Subfelder 
        return [[line[:5].rstrip(), None, None, line[8:]]]


    if len(sf_vals) == 0:
        # für den Fall, dass Leerzeilen vorkommen
        pass
    elif len(sf_vals) == 1:
        # wenn es nur ein Subfeld gibt
        out.append([line[:3], line[3:5], SUBFIELD_PREFIX + sf_vals[0][0], sf_vals[0][1:]])
    else:
        # falls es mehrere Subfelder gibt
        for i in range(len(sf_vals)):
            if i == 0:
                out.append([line[:3], line[3:5], SUBFIELD_PREFIX + sf_vals[0][0], sf_vals[i][1:]])
            else:
                out.append([None, None, SUBFIELD_PREFIX + sf_vals[i][0], sf_vals[i][1:]])
            
    return out



def write_to_xlsx(list_of_lines):   
    for i in range(len(list_of_lines)):
        ws.append(list_of_lines[i])

# Verarbeitung des Inputs in ein xlsx-File
for i in range(len(mrc_in) - 1):
    if len(mrc_in[i]) > 0:
        write_to_xlsx(process_line(mrc_in[i]))


def merge_empty(ws):
    for row in range(ws.max_row):
        cell = ws.cell(row=row + 1, column=1)
        if cell.value == None:
            ws.merge_cells(start_row=cell.row - 1, start_column=1,end_row=cell.row,end_column=2)
            print(cell.row)

            
merge_empty(ws)

print(ws["A1"])
wb.save(out_file_name)


